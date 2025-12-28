# js/management/commands/import_employees.py
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from core.models import Organization, Unit, JobRole, EmployeeProfile

from openpyxl import load_workbook
from datetime import datetime, date
from jdatetime import datetime as jdatetime_datetime  # ایمپورت شفاف‌تر برای IDE
import re

# --- Helpers ---------------------------------------------------------------

PERSIAN_DIGITS = "۰۱۲۳۴۵۶۷۸۹"
EN_DIGITS = "0123456789"
P2E = str.maketrans(PERSIAN_DIGITS, EN_DIGITS)

def normalize_text(s: str) -> str:
    """
    نرمال‌سازی متن: حذف نیم‌فاصله/فاصله نامرئی، تبدیل اعداد فارسی به انگلیسی، یک‌دست‌سازی فاصله‌ها.
    """
    if s is None:
        return ""
    s = str(s)
    # حذف نیم‌فاصله و NBSP
    s = s.replace("\u200c", " ").replace("\u00a0", " ")
    # تبدیل اعداد فارسی به انگلیسی
    s = s.translate(P2E)
    # یکدست‌سازی فاصله‌ها
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_iso_or_jalali_date(value):
    """
    پذیرش تاریخ:
      - آبجکت تاریخ/زمان اکسل (datetime/date)
      - رشته میلادی ISO مثل YYYY-MM-DD
      - رشته شمسی مثل 1402-01-15 یا 1402/01/15
    خروجی همیشه date میلادی است.
    """
    if not value:
        return None

    # اگر خود تاریخ/زمان پایتون/اکسل باشد
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    val = normalize_text(value)

    # میلادی ISO (YYYY-MM-DD)
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError:
        pass

    # شمسی با دو فرمت رایج
    for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
        try:
            jd = jdatetime_datetime.strptime(val, fmt)
            return jd.togregorian().date()
        except ValueError:
            continue

    raise CommandError(
        f"Invalid date format: {value!r} "
        f"(expected YYYY-MM-DD (میلادی) یا تاریخ شمسی مثل 1402-01-15)"
    )

# --- Command ---------------------------------------------------------------

class Command(BaseCommand):
    help = (
        "Import employees from an Excel (.xlsx). "
        "Username = personnel_code; password is unusable (SSO later). "
        "Flags: --autocreate-units --autocreate-roles --create-orgs"
    )

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str, help="Path to Excel file (.xlsx)")
        parser.add_argument("--sheet", type=str, default="Sheet1", help="Worksheet name (default: Sheet1)")
        parser.add_argument("--dry-run", action="store_true", help="Parse only; roll back all changes")
        parser.add_argument("--autocreate-units", action="store_true",
                            help="Auto-create Unit if not found in the given Organization.")
        parser.add_argument("--autocreate-roles", action="store_true",
                            help="Auto-create JobRole (is_active=True) if not found.")
        parser.add_argument("--create-orgs", action="store_true",
                            help="Auto-create Organization if not found (use with caution).")

    def handle(self, *args, **options):
        path = options["filepath"]
        sheet_name = options["sheet"]
        dry_run = options["dry_run"]
        autocreate_units = options["autocreate_units"]
        autocreate_roles = options["autocreate_roles"]
        create_orgs = options["create_orgs"]

        # Load workbook
        try:
            wb = load_workbook(filename=path, data_only=True)
        except FileNotFoundError as e:
            raise CommandError(f"Excel file not found: {e}")
        except Exception as e:
            raise CommandError(f"Cannot open Excel file: {e}")

        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Worksheet '{sheet_name}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet_name]

        # Headers
        headers_required = [
            "personnel_code", "first_name", "last_name", "organization",
            "unit", "job_role", "title", "hire_date",
        ]
        headers_optional = ["email", "direct_supervisor_personnel_code"]

        header_row = [normalize_text(c.value) for c in ws[1]]
        idx = {h: (header_row.index(h) if h in header_row else None) for h in headers_required + headers_optional}
        missing = [h for h in headers_required if idx[h] is None]
        if missing:
            raise CommandError(f"Missing required headers: {missing}. Found: {header_row}")

        rows = list(ws.iter_rows(min_row=2))

        created_users = 0
        upserted_profiles = 0
        errors = 0

        def cell(row, hname):
            i = idx[hname]
            return None if i is None else (row[i].value if i < len(row) else None)

        @transaction.atomic
        def import_one(row_num, row):
            nonlocal created_users, upserted_profiles

            # Read & normalize
            personnel_code = normalize_text(cell(row, "personnel_code"))
            first_name     = normalize_text(cell(row, "first_name"))
            last_name      = normalize_text(cell(row, "last_name"))
            org_name       = normalize_text(cell(row, "organization"))
            unit_name      = normalize_text(cell(row, "unit"))
            job_role_name  = normalize_text(cell(row, "job_role"))
            title          = normalize_text(cell(row, "title"))
            email          = normalize_text(cell(row, "email")) if idx["email"] is not None else ""
            sup_code       = normalize_text(cell(row, "direct_supervisor_personnel_code")) if idx["direct_supervisor_personnel_code"] is not None else ""
            hire_date_raw  = cell(row, "hire_date")
            hire_date      = parse_iso_or_jalali_date(hire_date_raw) if hire_date_raw else None

            # Basic validations
            if not personnel_code:
                raise CommandError(f"Row {row_num}: 'personnel_code' is required.")
            if not org_name:
                raise CommandError(f"Row {row_num}: 'organization' is required.")

            # Organization
            try:
                org = Organization.objects.get(name=org_name)
            except Organization.DoesNotExist:
                if create_orgs:
                    org = Organization.objects.create(name=org_name)
                else:
                    raise CommandError(
                        f"Row {row_num}: organization '{org_name}' not found. "
                        f"Create it first or use --create-orgs."
                    )

            # Unit
            unit = None
            if unit_name:
                try:
                    unit = Unit.objects.get(organization=org, name=unit_name)
                except Unit.DoesNotExist:
                    if autocreate_units:
                        unit = Unit.objects.create(organization=org,name=unit_name,supervision_policy="DEFAULT",)
                    else:
                        raise CommandError(
                            f"Row {row_num}: unit '{unit_name}' not found in organization '{org_name}'. "
                            f"Create it first or use --autocreate-units."
                        )

            # JobRole
            job_role = None
            if job_role_name:
                try:
                    job_role = JobRole.objects.get(name=job_role_name, is_active=True)
                except JobRole.DoesNotExist:
                    if autocreate_roles:
                        job_role = JobRole.objects.create(name=job_role_name, is_active=True)
                    else:
                        raise CommandError(
                            f"Row {row_num}: job_role '{job_role_name}' not found or inactive. "
                            f"Create it first or use --autocreate-roles."
                        )

            # User (username = personnel_code)
            user, created = User.objects.get_or_create(
                username=personnel_code,
                defaults={
                    "first_name": first_name,
                    "last_name": last_name,
                    "email": email,
                    "is_active": True,
                },
            )
            if created:
                user.set_unusable_password()  # SSO بعداً
                user.save()
                created_users += 1
            else:
                changed = False
                if first_name and user.first_name != first_name:
                    user.first_name = first_name; changed = True
                if last_name and user.last_name != last_name:
                    user.last_name = last_name; changed = True
                if email and user.email != email:
                    user.email = email; changed = True
                if changed:
                    user.save()

            # Optional: direct supervisor by personnel_code (باید قبلاً ایمپورت شده باشد)
            supervisor_user = None
            if sup_code:
                try:
                    supervisor_user = User.objects.get(username=sup_code)
                except User.DoesNotExist:
                    raise CommandError(
                        f"Row {row_num}: direct_supervisor personnel_code '{sup_code}' not found "
                        f"(import supervisor first or leave blank)."
                    )

            # Upsert EmployeeProfile
            defaults = {
                "organization": org,
                "unit": unit,
                "job_role": job_role,
                "personnel_code": personnel_code,
                "title": title or None,
                "hire_date": hire_date,
                "direct_supervisor": supervisor_user,
            }
            EmployeeProfile.objects.update_or_create(user=user, defaults=defaults)
            upserted_profiles += 1

        # Iterate
        for r_idx, row in enumerate(rows, start=2):
            if all((c.value is None or str(c.value).strip() == "") for c in row):
                continue
            try:
                if dry_run:
                    with transaction.atomic():
                        import_one(r_idx, row)
                        # rollback عمدی برای dry-run
                        raise transaction.TransactionManagementError("Dry-run rollback")
                else:
                    import_one(r_idx, row)
            except transaction.TransactionManagementError:
                # rollback مورد انتظار در dry-run
                pass
            except CommandError as ce:
                errors += 1
                self.stderr.write(self.style.ERROR(str(ce)))
            except ValueError as ve:
                errors += 1
                self.stderr.write(self.style.ERROR(f"Row {r_idx}: {ve}"))

        self.stdout.write(self.style.SUCCESS(
            f"Done. Users created: {created_users}, Profiles upserted: {upserted_profiles}, Errors: {errors}"
        ))
