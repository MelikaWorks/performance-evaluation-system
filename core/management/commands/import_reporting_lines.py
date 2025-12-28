from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.db import transaction
from openpyxl import load_workbook
from core.models import Organization, ReportingLine, EmployeeProfile

def norm(x):
    return "" if x is None else str(x).replace("\u200c"," ").replace("\u00a0"," ").strip()

class Command(BaseCommand):
    help = "Import reporting lines from XLSX (subordinate_personnel_code, supervisor_personnel_code, [organization])"

    def add_arguments(self, parser):
        parser.add_argument("filepath", type=str)
        parser.add_argument("--sheet", type=str, default="Sheet1")
        parser.add_argument("--org", type=str, help="Default organization name if column is missing")

    def handle(self, *args, **opts):
        path = opts["filepath"]; sheet = opts["sheet"]; org_default = opts.get("org")
        wb = load_workbook(filename=path, data_only=True)
        if sheet not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
        ws = wb[sheet]

        header = [norm(c.value) for c in ws[1]]
        def idx(name): return header.index(name) if name in header else None
        i_sub = idx("subordinate_personnel_code")
        i_sup = idx("supervisor_personnel_code")
        i_org = idx("organization")

        if i_sub is None or i_sup is None:
            raise CommandError("Headers required: subordinate_personnel_code, supervisor_personnel_code")

        created = updated = errors = 0

        with transaction.atomic():
            for r, row in enumerate(ws.iter_rows(min_row=2), start=2):
                sub_pc = norm(row[i_sub].value if i_sub is not None else "")
                sup_pc = norm(row[i_sup].value if i_sup is not None else "")
                org_name = norm(row[i_org].value if i_org is not None else org_default or "")

                if not sub_pc or not sup_pc or not org_name:
                    errors += 1; self.stderr.write(f"Row {r}: missing data"); continue

                try:
                    org = Organization.objects.get(name=org_name)
                except Organization.DoesNotExist:
                    errors += 1; self.stderr.write(f"Row {r}: org '{org_name}' not found"); continue

                try:
                    sub_user = User.objects.get(username=sub_pc)
                    sup_user = User.objects.get(username=sup_pc)
                except User.DoesNotExist:
                    errors += 1; self.stderr.write(f"Row {r}: user not found (sub:{sub_pc}, sup:{sup_pc})"); continue

                rl, is_new = ReportingLine.objects.update_or_create(
                    organization=org,
                    subordinate=sub_user,
                    defaults={"supervisor": sup_user},
                )
                if is_new: created += 1
                else: updated += 1

                # sync profile.direct_supervisor too (optional but handy)
                try:
                    p = EmployeeProfile.objects.get(user=sub_user, organization=org)
                    if p.direct_supervisor_id != sup_user.id:
                        p.direct_supervisor = sup_user
                        p.save(update_fields=["direct_supervisor"])
                except EmployeeProfile.DoesNotExist:
                    pass

        self.stdout.write(self.style.SUCCESS(
            f"Done. Created: {created}, Updated: {updated}, Errors: {errors}"
        ))
