import re
from pathlib import Path
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from core.models import FormTemplate, FormCriterion, FormOption, JobRole  # اگر JobRole داری

BOOL_TRUE = {"true","True","TRUE","1","yes","YES","required","Required","REQUIRED"}

def split_pipe(s):
    if s is None: return []
    parts = [str(x).strip() for x in str(s).split("|")]
    return [p for p in parts if p != ""]

def to_bool(v):
    if v is None: return False
    return str(v).strip() in BOOL_TRUE

class Command(BaseCommand):
    help = "Import HR Form Templates from XLSX files (sheets: «فرم», «سوالات»)."

    def add_arguments(self, parser):
        parser.add_argument("--files", nargs="+", help="XLSX files to import")
        parser.add_argument("--dir", help="Directory containing XLSX files")

    def handle(self, *args, **opts):
        files = []
        if opts.get("files"):
            files.extend([Path(p) for p in opts["files"]])
        if opts.get("dir"):
            files.extend(sorted(Path(opts["dir"]).glob("*.xlsx")))
        files = [p for p in files if p.exists()]
        if not files:
            raise CommandError("No input files found. Use --files or --dir.")

        for path in files:
            self.stdout.write(self.style.NOTICE(f"\n==> Importing {path.name}"))
            try:
                self.import_file(path)
            except Exception as e:
                raise CommandError(f"{path.name}: {e}")

    @transaction.atomic
    def import_file(self, path: Path):
        wb = load_workbook(filename=str(path), data_only=True)
        if "فرم" not in wb.sheetnames or "سوالات" not in wb.sheetnames:
            raise CommandError("Sheets «فرم» and «سوالات» are required.")

        # --- Sheet: فرم ---
        sh_form = wb["فرم"]
        # Expect header in row 1, data in row 2
        headers = [c.value for c in sh_form[1]]
        values  = [c.value for c in sh_form[2]]
        f = dict(zip(headers, values))

        code = (f.get("form_code") or "").strip()
        name = (f.get("form_name") or "").strip()
        if not code or not name:
            raise CommandError("form_code and form_name are required in «فرم» sheet.")

        description = f.get("description") or ""
        status = (f.get("status") or "Draft").strip() or "Draft"
        version = int(f.get("version") or 1)

        # flags (display only)
        show_emp_sig = to_bool(f.get("امضای کارمند"))
        show_mgr_sig = to_bool(f.get("امضای مدیر"))
        show_hr_sig  = to_bool(f.get("امضای منابع انسانی"))
        show_emp_cmt = to_bool(f.get("نظر کارمند"))
        show_goals   = to_bool(f.get("اهداف دوره بعد"))

        # role levels
        rl_raw = f.get("applies_to_role_levels") or ""
        role_levels = []
        if rl_raw:
            for piece in str(rl_raw).split(","):
                piece = piece.strip()
                if piece.isdigit():
                    role_levels.append(int(piece))

        # create/update templates (versioned)
        tmpl, created = FormTemplate.objects.get_or_create(code=code, version=version, defaults=dict(
            name=name, description=description, status="Draft",
            show_employee_signature=show_emp_sig, show_manager_signature=show_mgr_sig,
            show_hr_signature=show_hr_sig, show_employee_comment=show_emp_cmt,
            show_next_period_goals=show_goals, applies_to_role_levels=role_levels,
        ))
        if not created:
            # Only allowOverwrite if still Draft
            if tmpl.status != "Draft":
                raise CommandError(f"Template {code} v{version} is {tmpl.status}; cannot overwrite.")
            # update meta
            tmpl.name = name
            tmpl.description = description
            tmpl.show_employee_signature = show_emp_sig
            tmpl.show_manager_signature = show_mgr_sig
            tmpl.show_hr_signature = show_hr_sig
            tmpl.show_employee_comment = show_emp_cmt
            tmpl.show_next_period_goals = show_goals
            tmpl.applies_to_role_levels = role_levels
            tmpl.save()

        # map job roles by name (optional)
        jr_raw = f.get("applies_to_job_roles") or ""
        if jr_raw:
            names = [x.strip() for x in str(jr_raw).split(",") if x.strip()]
            found = []
            for nm in names:
                try:
                    jr = JobRole.objects.get(name__iexact=nm)
                    found.append(jr)
                except JobRole.DoesNotExist:
                    self.stdout.write(self.style.WARNING(f"JobRole not found: {nm}"))
            if found:
                tmpl.applies_to_jobroles.set(found)

        # wipe previous criteria/options for this version (idempotent import)
        tmpl.criteria.all().delete()

        # --- Sheet: سوالات ---
        sh_q = wb["سوالات"]
        q_headers = [c.value for c in sh_q[1]]

        def cell(row, key, default=None):
            if key not in q_headers: return default
            idx = q_headers.index(key)
            return sh_q[row][idx].value

        last_row = sh_q.max_row
        created_criteria = 0
        created_options = 0

        for r in range(2, last_row+1):
            order = cell(r, "ترتیب")
            title = cell(r, "معیار")
            desc  = cell(r, "شرح معیار")
            labels_s = cell(r, "درجات معیار")
            values_s = cell(r, "نمره معیارها")
            weight = cell(r, "وزن (اختیاری)") or 1

            if not order or not title or not labels_s or not values_s:
                # skip empty lines
                continue

            labels = split_pipe(labels_s)
            values = split_pipe(values_s)
            if len(labels) != len(values):
                raise CommandError(f"Row {r}: labels and values count mismatch.")

            # ensure numeric values
            try:
                nums = [float(v) for v in values]
            except:
                raise CommandError(f"Row {r}: non-numeric values in «نمره معیارها».")

            # auto-fix orientation if labels look best->worst but values ascending
            joined = "".join(labels)
            looks_best_to_worst = joined.startswith("بسیار خوب") or joined.startswith("بسیارخوب")
            ascending = all(nums[k] <= nums[k+1] for k in range(len(nums)-1))
            if looks_best_to_worst and ascending and len(nums) > 1:
                nums.reverse()

            crit = FormCriterion.objects.create(
                template=tmpl, order=int(order), title=str(title).strip(),
                description=str(desc or "").strip(), weight=weight or 1
            )
            created_criteria += 1

            for i, (lab, val) in enumerate(zip(labels, nums), start=1):
                FormOption.objects.create(
                    criterion=crit, order=i, label=str(lab).strip(), value=val
                )
                created_options += 1

        self.stdout.write(self.style.SUCCESS(
            f"{tmpl.code} v{tmpl.version}: criteria={created_criteria}, options={created_options}"
        ))
